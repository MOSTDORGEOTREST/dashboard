import os
import time
from dataclasses import dataclass
from datetime import date
from typing import Optional, Union, List
from dateutil import rrule
import xlrd
import openpyxl
import datetime
from config import configs
from db import Session, engine
from db import tables


def read_excel_cell(path: str, sheet_name: str, row: int, col: int) -> Union[str, float, None]:
    """
    Читает значение ячейки из Excel-файла (.xls или .xlsx)

    :param path: Путь к Excel-файлу
    :param sheet_name: Название листа
    :param row: Номер строки (0-индексация)
    :param col: Номер колонки (0-индексация)
    :return: Значение ячейки (str, float, None и т.п.)
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Файл не найден: {path}")

    if path.lower().endswith(".xls"):
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_name(sheet_name)
        return sheet.cell(row, col).value

    elif path.lower().endswith(".xlsx"):
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row + 1, column=col + 1)  # openpyxl — 1-индексация
        return cell.value

    else:
        raise ValueError("Формат файла должен быть .xls или .xlsx")

def find_excel_file(directory: str, year: str, month: str) -> str:
    """
    Возвращает путь к существующему Excel-файлу (.xlsx или .xls)

    :param directory: Базовая директория с Excel-файлами (str)
    :param year: Год, например '2025'
    :param month: Месяц, например '06'
    :return: Путь к найденному Excel-файлу (str)
    :raises FileNotFoundError: если файл не найден
    """
    base_filename = f"{month}.{year} - Учет офисного времени"
    year_dir = os.path.join(directory, year)

    candidates: List[str] = [
        os.path.join(year_dir, f"{base_filename}.xlsx"),
        os.path.join(year_dir, f"{base_filename}.xls"),
    ]

    for path in candidates:
        if os.path.exists(path):
            return path

    checked = "\n".join(candidates)
    raise FileNotFoundError(f"Excel-файл не найден. Проверены пути:\n{checked}")


@dataclass
class PrizeData:
    date: date
    value: float

def wait_for_excel_dir(possible_paths, retries=10, delay=3):
    for _ in range(retries):
        path = get_valid_excel_directory(possible_paths)
        if path:
            return path
        print("Директория не найдена, повтор через", delay, "сек")
        time.sleep(delay)
    return None

class PrizeParser:
    def __init__(self, excel_directory: str):
        self.excel_directory = excel_directory

        if not os.path.exists(self.excel_directory):
            print(self.excel_directory)
            raise FileNotFoundError("Отсутствует директория премий")

    def get_month_prize_from_excel(self, month: str, year: str) -> float:
        excel_file  = find_excel_file(
            directory=self.excel_directory,
            year=year,
            month=month
        )
        file_modified = os.path.getmtime(excel_file)
        file_modified_date = datetime.datetime.fromtimestamp(file_modified)
        print(f"Excel файл: {excel_file}")
        print(f"Последнее изменение файла: {file_modified_date}")

        prize_raw = read_excel_cell(
            path=excel_file,
            sheet_name='Итог',
            row=0,
            col=24
        )
        print(f"Значение из Excel ячейки [0,24]: {prize_raw}")
        if str(prize_raw).lower() in {"xxx", "ххх"}:
            raise ValueError("Неверная запись")

        return prize_raw

    def db_parse(self, current_date: date) -> Optional[PrizeData]:
        """
        Обновляет запись премии в БД, если она изменилась
        :param current_date: Дата, за которую проверяется премия
        :return: PrizeData если обновление произошло, иначе None
        """
        month = current_date.strftime('%m')
        year = "20" + current_date.strftime('%y')

        prize = self.get_month_prize_from_excel(
            month=month,
            year=year
        )
        prize_date = date(
            year=current_date.year,
            month=current_date.month,
            day=25
        )

        existing = self._db_get(prize_date)

        if existing:
            print(f"Дата: {prize_date}, Текущая премия в БД: {existing.value}, Новая премия: {prize}")
        else:
            print(f"Дата: {prize_date}, Запись не найдена в БД, Новая премия: {prize}")


        if existing is None or prize > existing.value:
            print(f"ОБНОВЛЯЕМ: existing={existing.value if existing else 'None'}, new={prize}")
            data = PrizeData(date=prize_date, value=prize)
            if existing:
                self._db_update(data)
            else:
                self._db_create(data)
            return data
        else:
            print(f"НЕ ОБНОВЛЯЕМ: existing={existing.value}, new={prize} (новое значение не больше существующего)")

        return None

    def daemon(self, general_update: bool = True):
        """
            Функция для непрерывного парсинга премий и сопутствующих данных
            :param excel_dir: Путь к директории с Excel-файлами
            :param delay_seconds: Задержка между итерациями. Если None — выполняется один раз.
            :param use_db: Нужно ли писать в БД
            """
        if general_update:
            try:
                tables.works.__table__.drop(engine)
                tables.works.__table__.create(engine)
            except Exception as err:
                print("⚠️ Ошибка при сбросе таблицы works:", err)

            prize_dates = [
                date(year=dt.year, month=dt.month, day=25)
                for dt in rrule.rrule(
                    rrule.MONTHLY, dtstart=date(2020, 5, 1), until=date.today()
                )
            ]

            try:
                for d in prize_dates:
                    self.db_parse(d)
                    print(f"✅ Премия за {d} успешно обновлена")
            except Exception as err:
                print("⚠️ Ошибка при обновлении премий:", err)
        else:
            today = date.today()
            date(year=today.year, month=today.month, day=25)

    def _db_get(self, prize_date: date) -> Optional[PrizeData]:
        session = Session()
        record = session.query(tables.prizes).filter_by(date=prize_date).first()
        session.close()

        return PrizeData(date=record.date, value=record.value) if record else None

    def _db_update(self, data: PrizeData) -> None:
        session = Session()
        prize = session.query(tables.prizes).filter_by(date=data.date).first()
        prize.date = data.date
        prize.value = data.value
        session.commit()
        session.close()

    def _db_create(self, data: PrizeData) -> None:
        session = Session()
        session.add(tables.prizes(**data.__dict__))
        session.commit()
        session.close()


def get_valid_excel_directory(possible_paths):
    for path in possible_paths:
        print(f"Проверка пути: {path}")
        if os.path.exists(path):
            print(f"Найден путь: {path}")
            return path
    return None

if __name__ == "__main__":
    possible_paths = [
        "/code/app/files/МДГТ - (Учет рабоч. времени, Отпуск, Даты рожд., телефоны, план работ, Исполнители)/УЧЕТ рабочего времени/",
    ]

    excel_directory = wait_for_excel_dir(possible_paths)

    if excel_directory is None:
        print("[ОШИБКА] Ни один путь к директории не найден. Работа остановлена.")
    else:
        try:
            parser = PrizeParser(excel_directory=excel_directory)

            month = date.today().strftime('%m')
            year = "20" + date.today().strftime('%y')

            record = parser.get_month_prize_from_excel(month='06', year='2023')
            print(record)

            parser.prize_daemon(general_update=True)

        except Exception as e:
            print(f"[ОШИБКА] Произошла ошибка при работе PrizeParser: {e}")
